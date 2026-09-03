import os
import json
import re
import threading
import time
from pathlib import Path
from datetime import datetime

import pyttsx3
from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, Response
from dotenv import load_dotenv
from google import genai
from google.genai import types

BASE_DIR = Path(__file__).resolve().parent


# ==================================================
# SETTINGS
# ==================================================

load_dotenv()

API_KEY = os.getenv("GEMINI_API_KEY")

if not API_KEY:
    raise RuntimeError(
        "GEMINI_API_KEY नहीं मिली। .env file check करें।"
    )

client = genai.Client(api_key=API_KEY)

app = FastAPI(title="Aila AI Assistant")

MEMORY_FILE = Path("memory.json")
REMINDER_FILE = Path("reminders.json")


# ==================================================
# ADVANCED MEMORY
# ==================================================

MEMORY_FILE = BASE_DIR / "memory.json"
REMINDER_FILE = BASE_DIR / "reminders.json"


def load_state():
    """Load long-term memories and chat history.

    Older Aila versions stored chat history as a plain list. That format is
    automatically migrated so existing chats are not lost.
    """
    if not MEMORY_FILE.exists():
        return {"memories": [], "chat_history": []}

    try:
        data = json.loads(MEMORY_FILE.read_text(encoding="utf-8"))

        if isinstance(data, list):
            # Old format: the whole file was chat history.
            return {"memories": [], "chat_history": data}

        if isinstance(data, dict):
            memories = data.get("memories", [])
            history = data.get("chat_history", [])
            if not isinstance(memories, list):
                memories = []
            if not isinstance(history, list):
                history = []
            return {"memories": memories, "chat_history": history}

    except Exception:
        pass

    return {"memories": [], "chat_history": []}


def save_state():
    MEMORY_FILE.write_text(
        json.dumps(
            {
                "memories": long_term_memories,
                "chat_history": chat_history,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


_state = load_state()
long_term_memories = _state["memories"]
chat_history = _state["chat_history"]


def normalize_memory(text):
    return re.sub(r"\s+", " ", str(text).strip())


def add_long_term_memory(text):
    text = normalize_memory(text)
    if not text:
        return False

    # Avoid exact duplicates.
    if any(normalize_memory(item).lower() == text.lower() for item in long_term_memories):
        return False

    long_term_memories.append(text)
    save_state()
    return True


def forget_long_term_memory(text):
    target = normalize_memory(text).lower()
    if not target:
        return 0

    old_count = len(long_term_memories)
    long_term_memories[:] = [
        item for item in long_term_memories
        if target not in normalize_memory(item).lower()
    ]
    removed = old_count - len(long_term_memories)

    if removed:
        save_state()

    return removed


def clear_all_long_term_memories():
    long_term_memories.clear()
    save_state()

# ==================================================
# REMINDERS
# ==================================================

def load_reminders():
    if not REMINDER_FILE.exists():
        return []

    try:
        data = json.loads(
            REMINDER_FILE.read_text(encoding="utf-8")
        )
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_reminders():
    REMINDER_FILE.write_text(
        json.dumps(
            reminders,
            ensure_ascii=False,
            indent=2
        ),
        encoding="utf-8"
    )


reminders = load_reminders()
pending_reminders = []
reminder_lock = threading.Lock()


# ==================================================
# COMPUTER VOICE
# ==================================================

def speak_reminder(message):
    try:
        engine = pyttsx3.init()
        engine.setProperty("rate", 150)
        engine.setProperty("volume", 1.0)

        # उपलब्ध Hindi/Indian voice चुनने की कोशिश
        try:
            voices = engine.getProperty("voices")

            for voice in voices:
                voice_text = (
                    str(getattr(voice, "name", ""))
                    + " "
                    + str(getattr(voice, "id", ""))
                ).lower()

                if (
                    "hindi" in voice_text
                    or "india" in voice_text
                    or "hi-in" in voice_text
                ):
                    engine.setProperty("voice", voice.id)
                    break

        except Exception:
            pass

        engine.say(message)
        engine.runAndWait()
        engine.stop()

    except Exception as error:
        print("❌ Voice Error:", error)


# ==================================================
# REMINDER CHECKER
# ==================================================

def reminder_checker():
    global reminders

    while True:
        try:
            reminders = load_reminders()

            now = datetime.now().strftime(
                "%Y-%m-%d %H:%M"
            )

            changed = False

            for reminder in reminders:
                if (
                    reminder.get("time") == now
                    and not reminder.get("done", False)
                ):
                    message = reminder.get(
                        "message",
                        "भाई जी, आपका reminder है।"
                    )

                    print("⏰ REMINDER:", message)

                    with reminder_lock:
                        pending_reminders.append(message)

                    # Computer speaker पर बोलेगा
                    speak_reminder(message)

                    reminder["done"] = True
                    changed = True

            if changed:
                save_reminders()

        except Exception as error:
            print("❌ Reminder Error:", error)

        time.sleep(20)


threading.Thread(
    target=reminder_checker,
    daemon=True
).start()


# ==================================================
# FILE / PDF CONTEXT
# ==================================================

active_file_name = None
active_file_text = ""
active_image_data = None
active_image_mime = None


def extract_uploaded_file(filename: str, data: bytes) -> str:
    """Extract readable text from common study/document files."""
    ext = Path(filename).suffix.lower()

    if ext in {".txt", ".md", ".py", ".js", ".html", ".css", ".json", ".csv"}:
        return data.decode("utf-8", errors="replace")

    if ext in {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}:
        return "__IMAGE_FILE__"

    if ext == ".pdf":
        from pypdf import PdfReader
        import io
        reader = PdfReader(io.BytesIO(data))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        return "\n\n".join(pages)

    if ext == ".docx":
        from docx import Document
        import io
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs)

    if ext == ".xlsx":
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets:
            rows.append(f"[Sheet: {ws.title}]")
            for row in ws.iter_rows(values_only=True):
                values = ["" if v is None else str(v) for v in row]
                rows.append(" | ".join(values))
        return "\n".join(rows)

    raise ValueError("यह file type अभी supported नहीं है। PDF, TXT, DOCX, XLSX, CSV, MD, PY, JS, HTML, CSS या JSON इस्तेमाल करें।")


# ==================================================
# PWA FILES
# ==================================================

@app.get("/manifest.json")
def manifest():
    manifest_data = {
        "name": "Aila AI Assistant",
        "short_name": "Aila",
        "description": "Aila - Your Personal AI Assistant",
        "id": "/",
        "start_url": "/",
        "scope": "/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#2196f3",
        "orientation": "portrait",
        "icons": [
            {
                "src": "/icons/icon-192.png",
                "sizes": "192x192",
                "type": "image/png"
            },
            {
                "src": "/icons/icon-512.png",
                "sizes": "512x512",
                "type": "image/png"
            }
        ]
    }
    return JSONResponse(content=manifest_data)


@app.get("/service-worker.js")
def service_worker():
    # Keep the service worker simple and reliable for local PWA installation.
    content = 'const CACHE_NAME = "aila-complete-v3";\n\nself.addEventListener("install", (event) => {\n    self.skipWaiting();\n});\n\nself.addEventListener("activate", (event) => {\n    event.waitUntil(self.clients.claim());\n});\n\nself.addEventListener("fetch", (event) => {\n    if (event.request.method !== "GET") return;\n    if (event.request.url.includes("/ask") || event.request.url.includes("/pending-reminders")) {\n        return;\n    }\n    event.respondWith(fetch(event.request));\n});\n'
    return Response(content=content, media_type="application/javascript", headers={"Cache-Control": "no-cache"})


@app.get("/icons/{filename}")
def app_icon(filename: str):
    return FileResponse(BASE_DIR / "icons" / filename, media_type="image/png")


APP_JS = r'''
if ("serviceWorker" in navigator) {
    window.addEventListener("load", function () {
        navigator.serviceWorker.register("/service-worker.js").catch(function (error) { console.log("PWA error:", error); });
    });
}
const input = document.getElementById("question");
const chat = document.getElementById("chat");
const sendBtn = document.getElementById("sendBtn");
const micBtn = document.getElementById("micBtn");
const voiceBtn = document.getElementById("voiceBtn");
const clearBtn = document.getElementById("clearBtn");
const memoryBtn = document.getElementById("memoryBtn");
const installBtn = document.getElementById("installBtn");
const webSearchBtn = document.getElementById("webSearchBtn");
const statusText = document.getElementById("status");
const fileInput = document.getElementById("fileInput");
const uploadBtn = document.getElementById("uploadBtn");
const removeFileBtn = document.getElementById("removeFileBtn");
const fileStatus = document.getElementById("fileStatus");

// HARD UI CLEANUP: remove duplicate controls left by older cached versions.
function cleanupDuplicateUI() {
    [".input-area", ".file-area", "#status", "#webSearchBtn", "#voiceBtn", "#clearBtn", "#memoryBtn", "#installBtn"].forEach(function (selector) {
        document.querySelectorAll(selector).forEach(function (node, index) {
            if (index > 0) node.remove();
        });
    });
}
cleanupDuplicateUI();
setTimeout(cleanupDuplicateUI, 100);
setTimeout(cleanupDuplicateUI, 500);
setTimeout(cleanupDuplicateUI, 1500);

let deferredInstallPrompt = null;
let voiceEnabled = false;
let webSearchEnabled = true;
function escapeHtml(text) { return String(text).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;").replace(/'/g,"&#039;"); }
function renderAilaText(text) {
    let s=escapeHtml(text).replace(/\r\n/g,"\n");
    s=s.replace(/^\s*([-*_])(?:\s*\1){2,}\s*$/gm,"");
    s=s.replace(/^\s*#{1,6}\s*(.+)$/gm,"<h3>$1</h3>");
    s=s.replace(/\*\*(.+?)\*\*/gs,"<strong>$1</strong>");
    s=s.replace(/__(.+?)__/gs,"<strong>$1</strong>");
    s=s.replace(/`([^`]+)`/g,"<code>$1</code>");
    s=s.replace(/^\s*[\*•]\s+(.+)$/gm,'<div class="list-item">• $1</div>');
    s=s.replace(/^\s*-\s+(.+)$/gm,'<div class="list-item">• $1</div>');
    s=s.replace(/\n{2,}/g,'<div class="para-gap"></div>');
    return s.replace(/\n/g,"<br>");
}
function speechText(text) { return String(text).replace(/[\*_#`]/g,"").replace(/[-]{3,}/g," ").replace(/\s+/g," ").trim(); }
function addMessage(text,type) {
    const wrapper=document.createElement("div"); wrapper.className="message "+type;
    const body=document.createElement("div"); body.className="message-text"; body.innerHTML=renderAilaText(text); wrapper.appendChild(body);
    if(type==="ai") { const button=document.createElement("button"); button.type="button"; button.className="message-speak"; button.innerText="🔊 सुनें"; button.addEventListener("click",function(){speak(text,true);}); wrapper.appendChild(button); }
    chat.appendChild(wrapper); chat.scrollTop=chat.scrollHeight;
}
function chooseHindiVoice(utterance) { const voices=speechSynthesis.getVoices(); const voice=voices.find(function(v){const lang=String(v.lang||"").toLowerCase();const name=String(v.name||"").toLowerCase();return lang.indexOf("hi-in")===0||name.indexOf("hindi")>=0;}); if(voice) utterance.voice=voice; }
function speak(text,force) {
    if(!window.speechSynthesis||!window.SpeechSynthesisUtterance) return;
    if(!voiceEnabled&&!force) return;
    const clean=speechText(text); if(!clean) return; window.speechSynthesis.cancel();
    const chunks=clean.match(/.{1,700}(?:\s|$)/g)||[clean]; let index=0;
    function nextChunk(){ if(index>=chunks.length)return; const u=new SpeechSynthesisUtterance(chunks[index++].trim()); u.lang="hi-IN"; u.rate=.92; u.pitch=1.02; chooseHindiVoice(u); u.onend=nextChunk; u.onerror=function(e){console.log("Speech error:",e.error);}; window.speechSynthesis.speak(u); }
    nextChunk();
}
webSearchBtn.addEventListener("click",function(){webSearchEnabled=!webSearchEnabled;webSearchBtn.innerText=webSearchEnabled?"🌐 Web Search: ON":"🌐 Web Search: OFF";statusText.innerText=webSearchEnabled?"Web Search चालू है।":"Web Search बंद है।";});
voiceBtn.addEventListener("click",function(){voiceEnabled=!voiceEnabled;if(voiceEnabled){voiceBtn.innerText="🔊 Voice ON ✓";speak("नमस्ते भाई जी, अब मैं आपको बोलकर भी समझा सकती हूँ।",true);}else{speechSynthesis.cancel();voiceBtn.innerText="🔇 Voice OFF";}});
async function askAI(){
    const question=input.value.trim(); if(!question)return; addMessage(question,"user"); input.value=""; statusText.innerText="Aila सोच रही है...";
    try{const response=await fetch("/ask?question="+encodeURIComponent(question)+"&web_search="+(webSearchEnabled?"true":"false")); const data=await response.json(); if(data.answer){addMessage(data.answer,"ai");const wantsVoice=/बोलकर|बोल के|आवाज़ से|आवाज से|सुनाकर|सुना दो|बोलो|voice|speak/i.test(question);if(wantsVoice)voiceEnabled=true;if(voiceEnabled||wantsVoice)speak(data.answer,true);}else addMessage("❌ जवाब नहीं मिला।","ai");}catch(error){addMessage("❌ Server से connection नहीं हो पाया।","ai");console.log(error);} statusText.innerText="";
}
sendBtn.addEventListener("click",askAI); input.addEventListener("keydown",function(e){if(e.key==="Enter")askAI();});
const SpeechRecognition=window.SpeechRecognition||window.webkitSpeechRecognition; let recognition=null;
if(SpeechRecognition){recognition=new SpeechRecognition();recognition.lang="hi-IN";recognition.continuous=false;recognition.interimResults=false;recognition.onstart=function(){micBtn.classList.add("listening");micBtn.innerText="🔴";statusText.innerText="🎙️ बोलिए...";};recognition.onresult=function(e){input.value=e.results[0][0].transcript;setTimeout(askAI,250);};recognition.onerror=function(e){micBtn.classList.remove("listening");micBtn.innerText="🎙️";statusText.innerText="❌ Microphone समस्या: "+e.error;};recognition.onend=function(){micBtn.classList.remove("listening");micBtn.innerText="🎙️";};micBtn.addEventListener("click",function(){try{recognition.start();}catch(error){console.log(error);}});}else micBtn.addEventListener("click",function(){alert("आपके browser में Voice Recognition उपलब्ध नहीं है। Chrome इस्तेमाल करें।");});
uploadBtn.addEventListener("click",async function(){const file=fileInput.files[0];if(!file){alert("पहले कोई PDF, फोटो या file चुनिए।");return;}uploadBtn.disabled=true;fileStatus.innerText="⏳ File पढ़ी जा रही है...";try{const response=await fetch("/upload-file",{method:"POST",headers:{"X-Filename":file.name},body:file});const data=await response.json();if(!response.ok||!data.success)throw new Error(data.error||"Upload failed");fileStatus.innerText="📄 "+data.message;removeFileBtn.disabled=false;}catch(error){fileStatus.innerText="❌ "+error.message;}finally{uploadBtn.disabled=false;}});
removeFileBtn.addEventListener("click",async function(){try{const response=await fetch("/remove-file",{method:"POST"});const data=await response.json();fileStatus.innerText=data.message||"File हटा दी गई है।";fileInput.value="";removeFileBtn.disabled=true;}catch(error){fileStatus.innerText="❌ File हटाने में समस्या हुई।";}});
memoryBtn.addEventListener("click",async function(){try{const response=await fetch("/memory");const data=await response.json();if(!data.memories||data.memories.length===0){alert("🧠 Aila की long-term memory अभी खाली है।");return;}const text=data.memories.map(function(item,index){return(index+1)+". "+item;}).join("\n");const remove=prompt("🧠 Aila की यादें:\n\n"+text+"\n\nहटाने के लिए memory का कोई शब्द लिखें।");if(remove&&remove.trim()){await fetch("/memory/forget?text="+encodeURIComponent(remove.trim()),{method:"POST"});alert("Memory हटाने की कोशिश पूरी हो गई।");}}catch(error){alert("Memory panel नहीं खुल पाया।");}});
async function checkReminders(){try{const response=await fetch("/pending-reminders");const data=await response.json();if(data.reminders&&data.reminders.length>0)data.reminders.forEach(function(message){addMessage("⏰ "+message,"ai");speak("भाई जी, "+message,true);});}catch(error){console.log(error);}} setInterval(checkReminders,5000);checkReminders();
clearBtn.addEventListener("click",async function(){try{await fetch("/clear",{method:"POST"});chat.innerHTML='<div class="message ai"><div class="message-text"><h3>नमस्ते! 👋 मैं Aila हूँ।</h3>Chat clear हो गई है। अब नया सवाल पूछिए।</div></div>';}catch(error){console.log(error);}});
window.addEventListener("beforeinstallprompt",function(event){event.preventDefault();deferredInstallPrompt=event;installBtn.style.display="block";installBtn.innerText="📲 Install Aila";});window.addEventListener("appinstalled",function(){deferredInstallPrompt=null;installBtn.innerText="✅ Aila Installed";installBtn.disabled=true;});installBtn.addEventListener("click",async function(){if(!deferredInstallPrompt){alert("Chrome ने अभी Aila के लिए installation prompt उपलब्ध नहीं कराया है।");return;}deferredInstallPrompt.prompt();const choice=await deferredInstallPrompt.userChoice;if(choice.outcome==="accepted")installBtn.innerText="✅ Aila Installed";deferredInstallPrompt=null;});
'''



@app.get("/app.js")
def app_js():
    return Response(
        content=APP_JS,
        media_type="application/javascript",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )

# ==================================================
# FILE UPLOAD API
# ==================================================

@app.post("/upload-file")
async def upload_file(request: Request):
    global active_file_name, active_file_text, active_image_data, active_image_mime

    filename = Path(request.headers.get("x-filename", "uploaded_file")).name
    data = await request.body()
    if not data:
        return JSONResponse({"success": False, "error": "File खाली है।"}, status_code=400)
    if len(data) > 15 * 1024 * 1024:
        return JSONResponse({"success": False, "error": "File 15 MB से छोटी रखें।"}, status_code=413)

    try:
        ext = Path(filename).suffix.lower()
        active_file_name = None
        active_file_text = ""
        active_image_data = None
        active_image_mime = None

        if ext in {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}:
            mime_map = {".jpg":"image/jpeg", ".jpeg":"image/jpeg", ".png":"image/png", ".webp":"image/webp", ".gif":"image/gif", ".bmp":"image/bmp"}
            active_file_name = filename
            active_image_data = data
            active_image_mime = mime_map[ext]
            return {"success": True, "filename": filename, "characters": 0, "is_image": True, "message": f"{filename} upload हो गई है। अब इस फोटो के बारे में सवाल पूछ सकते हैं।"}

        text = extract_uploaded_file(filename, data).strip()
        if not text:
            return JSONResponse({"success": False, "error": "इस file से readable text नहीं मिला।"}, status_code=400)
        active_file_name = filename
        active_file_text = text[:60000]
        return {"success": True, "filename": filename, "characters": len(text), "is_image": False, "message": f"{filename} upload हो गई है। अब इसके बारे में सवाल पूछ सकते हैं।"}
    except Exception as error:
        return JSONResponse({"success": False, "error": str(error)}, status_code=400)

@app.post("/remove-file")
def remove_file():
    global active_file_name, active_file_text, active_image_data, active_image_mime
    active_file_name = None
    active_file_text = ""
    active_image_data = None
    active_image_mime = None
    return {"success": True, "message": "Active file हटा दी गई है।"}


# HOME PAGE
# ==================================================

@app.get("/", response_class=HTMLResponse)
def home():
    return """
<!DOCTYPE html>
<html lang="hi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#2196f3">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<meta name="apple-mobile-web-app-title" content="Aila">

<title>Aila AI Assistant</title>

<style>
* {
    box-sizing: border-box;
}

body {
    margin: 0;
    font-family: Arial, "Noto Sans Devanagari", sans-serif;
    background: #eef3f8;
}

.header {
    background: #2563eb;
    color: white;
    padding: 18px;
    text-align: center;
    font-size: 25px;
    font-weight: bold;
}

.container {
    max-width: 980px;
    margin: 25px auto;
    background: white;
    border-radius: 18px;
    padding: 25px;
    box-shadow: 0 8px 30px rgba(0,0,0,0.10);
}

.chat {
    height: 480px;
    overflow-y: auto;
    padding: 10px;
}

.message {
    padding: 16px 18px;
    margin: 12px 0;
    border-radius: 16px;
    line-height: 1.75;
    white-space: normal;
    font-size: 17px;
    box-shadow: 0 2px 10px rgba(15,23,42,0.04);
}

.message-text h3 { margin: 0 0 10px; font-size: 20px; font-weight: 800; }
.message-text strong { font-weight: 800; }
.message-text code { background: #e2e8f0; padding: 2px 6px; border-radius: 6px; }
.list-item { margin: 5px 0; padding-left: 4px; }
.para-gap { height: 8px; }
.message-speak { margin-top: 12px; padding: 7px 12px; border-radius: 9px; font-size: 13px; background: #e2e8f0; color: #0f172a; }

.ai {
    background: #f1f5f9;
    text-align: left;
}

.user {
    background: #dbeafe;
    text-align: right;
}

.input-area {
    display: flex;
    gap: 10px;
    margin-top: 15px;
    align-items: center;
    position: sticky;
    bottom: 0;
    background: rgba(255,255,255,0.96);
    padding: 10px 0 4px;
    backdrop-filter: blur(8px);
}

#question {
    flex: 1;
    padding: 15px;
    border: 1px solid #cbd5e1;
    border-radius: 12px;
    font-size: 16px;
    outline: none;
}

button {
    border: none;
    border-radius: 12px;
    padding: 12px 18px;
    font-size: 16px;
    cursor: pointer;
}

#sendBtn {
    background: #2563eb;
    color: white;
}

#micBtn {
    background: #111827;
    color: white;
    min-width: 60px;
}

#micBtn.listening {
    background: #dc2626;
}

#webSearchBtn { width: 100%; margin-top: 12px; background: #0ea5e9; color: white; }

#voiceBtn {
    width: 100%;
    margin-top: 12px;
    background: #16a34a;
    color: white;
}

#clearBtn {
    width: 100%;
    margin-top: 12px;
    background: #ef4444;
    color: white;
}

#memoryBtn {
    width: 100%;
    margin-top: 12px;
    background: #0f766e;
    color: white;
}

#installBtn {
    width: 100%;
    margin-top: 12px;
    background: #7c3aed;
    color: white;
    display: none;
}


.file-area {
    margin-top: 14px;
    padding: 12px;
    border: 1px dashed #94a3b8;
    border-radius: 12px;
    background: #f8fafc;
}

.file-area input {
    max-width: 100%;
    margin-bottom: 8px;
}

#uploadBtn, #removeFileBtn {
    margin-right: 8px;
    margin-top: 4px;
}

#uploadBtn { background: #334155; color: white; }
#removeFileBtn { background: #64748b; color: white; }
#removeFileBtn:disabled { opacity: .5; cursor: not-allowed; }
#fileStatus { margin-top: 8px; color: #475569; font-size: 14px; }
#status {
    text-align: center;
    color: #64748b;
    margin-top: 10px;
    min-height: 22px;
}

@media (max-width: 650px) {
    .container {
        margin: 10px;
        padding: 15px;
    }

    .input-area {
        flex-wrap: wrap;
    }

    #question {
        width: 100%;
        flex-basis: 100%;
    }
}
</style>
</head>

<body>

<div class="header">
    🤖 Aila AI Assistant
    <div style="font-size:13px;font-weight:400;opacity:.9;margin-top:4px;">आपकी Personal AI Assistant</div>
</div>

<div class="container">

    <div id="chat" class="chat">
        <div class="message ai">
            नमस्ते! 👋 मैं Aila हूँ।<br>
            आप मुझसे कोई भी सवाल पूछ सकते हैं।
        </div>
    </div>

    <div class="file-area">
        <input id="fileInput" type="file" accept=".pdf,.txt,.md,.py,.js,.html,.css,.json,.csv,.docx,.xlsx,.jpg,.jpeg,.png,.webp,.gif,.bmp">
        <button id="uploadBtn" type="button">📁 Upload File</button>
        <button id="removeFileBtn" type="button" disabled>✖️ Remove File</button>
        <div id="fileStatus">कोई file upload नहीं है।</div>
    </div>

    <div id="status"></div>

    <div class="input-area">

        <input
            id="question"
            type="text"
            placeholder="अपना सवाल लिखें..."
            autocomplete="off"
        >

        <button id="micBtn" type="button">
            🎙️
        </button>

        <button id="sendBtn" type="button">
            Send
        </button>

    </div>

    <button id="webSearchBtn" type="button">
        🌐 Web Search: ON
    </button>

    <button id="voiceBtn" type="button">
        🔊 Voice ON
    </button>

    <button id="clearBtn" type="button">
        🗑️ Clear Chat
    </button>

    <button id="memoryBtn" type="button">
        🧠 My Memory
    </button>

    <button id="installBtn" type="button">
        📲 Install Aila
    </button>

</div>


<script src="/app.js"></script>

</body>
</html>
"""


# ==================================================
# PENDING REMINDERS
# ==================================================

@app.get("/pending-reminders")
def get_pending_reminders():

    global pending_reminders

    with reminder_lock:

        data = pending_reminders.copy()
        pending_reminders.clear()

    return {
        "reminders": data
    }


# ==================================================
# MANUAL ADD REMINDER
# ==================================================

@app.post("/add-reminder")
def add_reminder(time: str, message: str):

    reminders.append({
        "time": time,
        "message": message,
        "done": False
    })

    save_reminders()

    return {
        "success": True,
        "message": "Reminder successfully added."
    }


# ==================================================
# MEMORY API
# ==================================================

@app.get("/memory")
def get_memory():
    return {
        "memories": long_term_memories.copy()
    }


@app.post("/memory/forget")
def forget_memory(text: str):
    removed = forget_long_term_memory(text)
    return {
        "success": True,
        "removed": removed
    }


@app.post("/memory/clear")
def clear_memory():
    clear_all_long_term_memories()
    return {
        "success": True,
        "message": "Long-term memory cleared"
    }


# ==================================================
# ASK AI
# ==================================================

@app.get("/ask")
def ask(question: str, web_search: bool = True):
    global active_image_data, active_image_mime
    question = question.strip()
    if not question:
        return {"answer": "कृपया कोई सवाल लिखिए।"}

    conversation = "".join("User: " + str(i.get("question", "")) + "\nAila: " + str(i.get("answer", "")) + "\n\n" for i in chat_history[-10:])
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    memory_context = "\n".join(f"- {item}" for item in long_term_memories[-50:]) or "(अभी कोई long-term memory नहीं है)"
    prompt = f"""
तुम Aila नाम की intelligent AI Assistant हो।
वर्तमान तारीख और समय: {current_time}
यूज़र हिंदी में पूछे तो हिंदी में जवाब दो। जवाब सरल, स्पष्ट, उपयोगी और professional हो।

FORMAT नियम:
- साफ headings के लिए # या ## इस्तेमाल करो।
- जरूरी शब्दों को **bold** करो।
- lists के लिए bullet points इस्तेमाल करो।
- अनावश्यक * या --- मत लगाओ।

REMINDER SYSTEM:
यदि यूज़र reminder/alarm चाहता है और समय स्पष्ट है, तो:
REMINDER:
TIME: YYYY-MM-DD HH:MM
MESSAGE: reminder का संदेश
यदि reminder नहीं मांगा गया है तो यह format मत दो।

LONG-TERM MEMORY:
{memory_context}
यदि यूज़र किसी बात को याद रखने को कहे: MEMORY_SAVE: याद रखने वाली बात
यदि भूलने को कहे: MEMORY_FORGET: भूलने वाली बात

CHAT HISTORY:
{conversation}

FILE CONTEXT:
Active file: {active_file_name or "कोई file upload नहीं है"}
{active_file_text[:60000] if active_file_text else "(कोई text file context नहीं)"}

USER QUESTION:
{question}
"""
    models = ["gemini-3-flash-preview", "gemini-3.5-flash-lite"]

    def call_model(model, use_search):
        config = types.GenerateContentConfig(tools=[types.Tool(google_search=types.GoogleSearch())]) if use_search else None
        contents = prompt
        if active_image_data and active_image_mime:
            contents = [prompt, types.Part.from_bytes(data=active_image_data, mime_type=active_image_mime)]
        return client.models.generate_content(model=model, contents=contents, config=config)

    response = None
    search_used = False
    last_error = None
    if web_search:
        for model in models:
            try:
                response = call_model(model, True)
                search_used = True
                break
            except Exception as error:
                last_error = error
    if response is None:
        for model in models:
            try:
                response = call_model(model, False)
                search_used = False
                break
            except Exception as error:
                last_error = error
    if response is None:
        return {"question": question, "answer": "❌ अभी AI server से जवाब नहीं मिल पाया। थोड़ी देर बाद फिर कोशिश करें।", "error": str(last_error)}

    answer = (response.text or "मुझे जवाब नहीं मिला।").strip()
    if search_used:
        sources = []
        try:
            metadata = response.candidates[0].grounding_metadata
            for chunk in (getattr(metadata, "grounding_chunks", None) or []):
                web = getattr(chunk, "web", None)
                uri = getattr(web, "uri", None) if web else None
                title = getattr(web, "title", None) if web else None
                if uri and uri not in [x[1] for x in sources]:
                    sources.append((title or uri, uri))
        except Exception:
            pass
        if sources:
            answer += "\n\n## 🌐 स्रोत\n" + "\n".join(f"- {title} — {uri}" for title, uri in sources[:8])

    save_match = re.search(r"MEMORY_SAVE:\s*(.+)", answer, re.IGNORECASE)
    if save_match:
        saved = save_match.group(1).strip().split("\n")[0].strip()
        if saved:
            add_long_term_memory(saved)
            answer = "ठीक है भाई जी। 🧠 मैंने इसे अपनी long-term memory में याद रख लिया है।"
    forget_match = re.search(r"MEMORY_FORGET:\s*(.+)", answer, re.IGNORECASE)
    if forget_match:
        target = forget_match.group(1).strip().split("\n")[0].strip()
        removed = forget_long_term_memory(target)
        answer = "ठीक है भाई जी। 🧠 मैंने वह बात अपनी long-term memory से हटा दी है।" if removed else "ठीक है भाई जी। 🧠 ऐसी memory नहीं मिली।"

    reminder_match = re.search(r"REMINDER:\s*TIME:\s*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})\s*MESSAGE:\s*(.+)", answer, re.IGNORECASE | re.DOTALL)
    if reminder_match:
        reminder_time = reminder_match.group(1).strip()
        reminder_message = reminder_match.group(2).strip().split("\n\n")[0].strip()
        reminders.append({"time": reminder_time, "message": reminder_message, "done": False})
        save_reminders()
        answer = f"ठीक है भाई जी। ⏰\nमैंने आपका reminder {reminder_time} के लिए लगा दिया है।"

    chat_history.append({"question": question, "answer": answer})
    save_state()
    return {"question": question, "answer": answer, "web_search_used": search_used}


# ==================================================
# CLEAR CHAT
# ==================================================

@app.post("/clear")
def clear_chat():

    chat_history.clear()

    save_state()

    return {
        "message": "Chat history cleared"
    }
